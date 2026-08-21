#!/usr/bin/env python3
"""Read 500kata.xlsx and seed flashcards.db with cards + empty study_progress."""

import sqlite3
import openpyxl
import os
import sys
from datetime import datetime

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "Downloads", "500kata.xlsx")
DB_PATH = os.path.join(os.path.dirname(__file__), "flashcards.db")

def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else XLSX_PATH
    print(f"📖 Reading: {xlsx}")
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["TOEFL Detailed Word List"]

    # Create DB
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print("🗑  Removed old DB")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Schema
    cur.executescript("""
        CREATE TABLE cards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lesson INTEGER NOT NULL,
            word TEXT NOT NULL,
            pos TEXT,
            synonym TEXT,
            definition TEXT,
            example1 TEXT,
            example2 TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE study_progress (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            card_id INTEGER NOT NULL,
            direction TEXT NOT NULL CHECK(direction IN ('kw_to_syn', 'syn_to_kw')),
            pile TEXT NOT NULL DEFAULT 'unmastered' CHECK(pile IN ('mastered', 'unmastered')),
            interval INTEGER DEFAULT 1,
            ease_factor REAL DEFAULT 2.5,
            repetitions INTEGER DEFAULT 0,
            next_review TEXT,
            last_reviewed TEXT,
            FOREIGN KEY (card_id) REFERENCES cards(id) ON DELETE CASCADE,
            UNIQUE(card_id, direction)
        );

        CREATE INDEX idx_cards_lesson ON cards(lesson);
        CREATE INDEX idx_progress_pile ON study_progress(pile);
        CREATE INDEX idx_progress_review ON study_progress(next_review);
    """)

    # Import data
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        lesson, word, pos, synonym, definition, ex1, ex2 = row
        if not word:
            continue

        # Clean placeholder text
        def clean(val):
            if val and "[Study Input Needed" in str(val):
                return ""
            return val or ""

        cur.execute(
            "INSERT INTO cards (lesson, word, pos, synonym, definition, example1, example2) VALUES (?,?,?,?,?,?,?)",
            (lesson, word.strip(), pos, clean(synonym), clean(definition), clean(ex1), clean(ex2))
        )
        card_id = cur.lastrowid

        # Create study_progress for both directions
        now = datetime.utcnow().isoformat()
        for direction in ("kw_to_syn", "syn_to_kw"):
            cur.execute(
                "INSERT INTO study_progress (card_id, direction, pile, next_review) VALUES (?,?,?,?)",
                (card_id, direction, "unmastered", now)
            )
        count += 1

    conn.commit()
    conn.close()
    print(f"✅ Seeded {count} cards with {count * 2} study_progress entries")
    print(f"📦 Database: {DB_PATH}")

if __name__ == "__main__":
    main()
